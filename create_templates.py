import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "Title", "Start Date", "End Date", "Start Time", "End Time",
    "All Day", "Location", "Description", "Category", "URL",
    "Status", "Priority", "Transparency", "Classification",
    "Reminder", "Recurrence", "Recurrence End Date", "Recurrence Count",
    "Recurrence Interval", "Organizer", "Required Attendees",
    "Optional Attendees", "Duration"
]

COL_WIDTHS = [22, 14, 14, 12, 12, 8, 25, 30, 15, 25,
              12, 10, 14, 14, 12, 14, 18, 16, 18, 25, 30, 30, 14]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")

ALLDAY = '"Yes,No"'
STATUSES = '"CONFIRMED,TENTATIVE,CANCELLED"'
TRANSP = '"OPAQUE,TRANSPARENT"'
CLASS = '"PUBLIC,PRIVATE,CONFIDENTIAL"'
RECURRENCES = '"Daily,Weekly,Fortnightly,Monthly,Yearly,Weekday"'
REMINDERS = '"5,10,15,30,60,120,1440,2880"'

def style_header_row(ws):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    ws.freeze_panes = "A2"

def add_validations(ws, max_row=100):
    validations = [
        (ALLDAY, "F"),
        (STATUSES, "K"),
        (TRANSP, "M"),
        (CLASS, "N"),
        (RECURRENCES, "P"),
    ]
    for formula, col in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.add(f"{col}2:{col}{max_row}")
        ws.add_data_validation(dv)

def add_notes_sheet(wb):
    ns = wb.create_sheet("Notes")
    ns.sheet_properties.tabColor = "FFC000"
    ns.column_dimensions['A'].width = 22
    ns.column_dimensions['B'].width = 55
    ns.column_dimensions['C'].width = 45
    ns.column_dimensions['D'].width = 12

    note_headers = ['Column', 'Description', 'Accepted Values / Examples', 'Required']
    for col, h in enumerate(note_headers, 1):
        cell = ns.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    notes_data = [
        ("Title", "Event title/summary", "Team Meeting, Birthday Party", "Yes"),
        ("Start Date", "Event start date", "29/01/2026, 2026-01-29, 01/29/2026", "Yes"),
        ("End Date", "Event end date (optional if Duration used)", "30/01/2026", "No"),
        ("Start Time", "Start time (leave blank for all-day)", "09:00, 9:30 AM, 14:00", "No"),
        ("End Time", "End time", "17:00, 5:00 PM", "No"),
        ("All Day", "Whether all-day event", "Yes, No", "No"),
        ("Location", "Venue or address", "Conference Room B, 10 Downing St", "No"),
        ("Description", "Event details/notes", "Free text", "No"),
        ("Category", "Event categories", "Work, Personal, Meeting", "No"),
        ("URL", "Related web link", "https://example.com", "No"),
        ("Status", "Event status", "CONFIRMED, TENTATIVE, CANCELLED", "No"),
        ("Priority", "Importance (0=undefined, 1=highest, 9=lowest)", "0-9", "No"),
        ("Transparency", "Free/busy indicator", "OPAQUE (busy), TRANSPARENT (free)", "No"),
        ("Classification", "Visibility/access level", "PUBLIC, PRIVATE, CONFIDENTIAL", "No"),
        ("Reminder", "Minutes before event to alert", "15, 60, 1440, or '2 hours', '1 day'", "No"),
        ("Recurrence", "Repeat pattern", "Daily, Weekly, Monthly, Yearly, Fortnightly, Weekday", "No"),
        ("Recurrence End Date", "When recurring events stop", "31/12/2026", "No"),
        ("Recurrence Count", "Number of occurrences", "10, 52", "No"),
        ("Recurrence Interval", "Gap between occurrences", "2 (=every 2 weeks/months)", "No"),
        ("Organizer", "Meeting organizer email", "john@example.com", "No"),
        ("Required Attendees", "Required attendee emails (;-separated)", "a@ex.com; b@ex.com", "No"),
        ("Optional Attendees", "Optional attendee emails (;-separated)", "c@ex.com; d@ex.com", "No"),
        ("Duration", "Event length (alternative to End Date/Time)", "1 hour, 30 minutes, 1.5 hours", "No"),
    ]

    body_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    req_font = Font(name="Arial", size=10, bold=True, color="CC0000")

    for row, (col_name, desc, examples, req) in enumerate(notes_data, 2):
        for c, val in enumerate([col_name, desc, examples, req], 1):
            cell = ns.cell(row=row, column=c, value=val)
            cell.font = body_font
            cell.border = THIN_BORDER
            if c == 1:
                cell.font = bold_font
            if c == 4 and req == "Yes":
                cell.font = req_font
    ns.freeze_panes = "A2"

    # Tips section
    tip_row = len(notes_data) + 3
    ns.cell(row=tip_row, column=1, value="Tips:").font = Font(name="Arial", bold=True, size=10)
    tips = [
        "Only Title and Start Date are required. All other fields are optional.",
        "Select your date format on Page 1 of CSV2ICS to match your data.",
        "Default reminders can also be set on the Export page.",
        "Save as CSV (UTF-8) before importing. TSV (tab-delimited) is also supported.",
        "Attendee emails should be separated by semicolons.",
        "Duration is used instead of End Date/Time when End Date is not provided.",
        "Recurrence Count and End Date are mutually exclusive (RFC 5545).",
    ]
    for i, tip in enumerate(tips):
        ns.cell(row=tip_row + 1 + i, column=1, value=f"\u2022 {tip}").font = Font(name="Arial", size=9, italic=True, color="666666")


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"
    ws.sheet_properties.tabColor = "2F5496"
    style_header_row(ws)
    add_validations(ws)
    for row in range(2, 21):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_FILL
    add_notes_sheet(wb)
    wb.save(r"C:\Users\simcr\OneDrive\Projects\CSV2ICS\CSV2ICS_Template.xlsx")
    print("Template saved")


def create_sample():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"
    ws.sheet_properties.tabColor = "2F5496"
    style_header_row(ws)
    add_validations(ws, max_row=20)

    # 15 sample events demonstrating all features
    # Title, StartDate, EndDate, StartTime, EndTime, AllDay, Location, Description,
    # Category, URL, Status, Priority, Transparency, Classification,
    # Reminder, Recurrence, RecurEndDate, RecurCount, RecurInterval,
    # Organizer, ReqAttendees, OptAttendees, Duration
    samples = [
        ["Team Stand-up", "27/03/2026", "", "09:00", "09:15", "", "Board Room",
         "Daily sync meeting", "Work", "", "CONFIRMED", "1", "OPAQUE", "PUBLIC",
         "5", "Weekday", "30/06/2026", "", "",
         "manager@company.com", "dev1@company.com; dev2@company.com", "pm@company.com", ""],
        ["Pay Day", "28/03/2026", "", "", "", "Yes", "",
         "Monthly salary payment", "Finance", "", "CONFIRMED", "", "", "PRIVATE",
         "1440", "Monthly", "", "12", "",
         "", "", "", ""],
        ["Annual Leave", "06/04/2026", "10/04/2026", "", "", "Yes", "",
         "Easter holiday", "Personal", "", "CONFIRMED", "", "TRANSPARENT", "PRIVATE",
         "", "", "", "", "",
         "", "", "", ""],
        ["Project Deadline", "2026-04-15", "", "17:00", "", "", "Head Office",
         "Q1 deliverables due", "Work", "https://project.example.com", "TENTATIVE", "2", "OPAQUE", "CONFIDENTIAL",
         "60", "", "", "", "",
         "", "", "", ""],
        ["Dentist Appointment", "20/04/2026", "", "14:30", "", "", "42 Harley Street, London",
         "Six month check-up", "Health", "", "CONFIRMED", "3", "OPAQUE", "PRIVATE",
         "120", "", "", "", "",
         "", "", "", "1 hour"],
        ["Birthday Party", "25/04/2026", "", "18:00", "23:00", "", "The Red Lion, Manchester",
         "Surprise party for Sarah", "Social", "", "CONFIRMED", "", "", "PUBLIC",
         "1440", "", "", "", "",
         "", "friend1@email.com; friend2@email.com", "friend3@email.com", ""],
        ["Sprint Planning", "01/04/2026", "", "10:00", "", "", "Meeting Room 3",
         "Fortnightly sprint planning session", "Work", "", "CONFIRMED", "1", "OPAQUE", "PUBLIC",
         "15", "Fortnightly", "30/09/2026", "", "",
         "scrum@company.com", "team@company.com", "", "2 hours"],
        ["Gym Class", "30/03/2026", "", "07:00", "08:00", "", "PureGym, Birmingham",
         "Spin class", "Health", "", "CONFIRMED", "", "OPAQUE", "PRIVATE",
         "30", "Weekly", "", "52", "",
         "", "", "", ""],
        ["Board Meeting", "2026-05-01", "", "09:00", "12:00", "", "HQ Conference Suite, London",
         "Quarterly board review", "Work", "", "CONFIRMED", "1", "OPAQUE", "CONFIDENTIAL",
         "1440", "Yearly", "", "", "",
         "ceo@company.com", "board@company.com", "secretary@company.com", ""],
        ["Car MOT", "15/05/2026", "", "08:30", "", "", "Kwik Fit, Leeds",
         "Annual MOT test", "Personal", "", "CONFIRMED", "2", "OPAQUE", "PRIVATE",
         "1440", "Yearly", "", "", "",
         "", "", "", ""],
        ["Team Lunch", "03/04/2026", "", "12:30", "", "", "Wagamama, Bristol",
         "Monthly team social", "Social", "", "CONFIRMED", "", "TRANSPARENT", "PUBLIC",
         "60", "Monthly", "31/12/2026", "", "",
         "", "team@company.com", "", "1.5 hours"],
        ["Webinar: Cloud Security", "10/04/2026", "", "14:00", "15:30", "", "",
         "Online security best practices", "Training", "https://webinar.example.com/cloud", "CONFIRMED", "", "OPAQUE", "PUBLIC",
         "15", "", "", "", "",
         "", "", "", ""],
        ["Bank Holiday", "04/05/2026", "", "", "", "Yes", "",
         "Early May bank holiday", "Public Holiday", "", "CONFIRMED", "", "TRANSPARENT", "PUBLIC",
         "", "", "", "", "",
         "", "", "", ""],
        ["1:1 with Manager", "02/04/2026", "", "11:00", "", "", "Office 204",
         "Weekly catch-up", "Work", "", "CONFIRMED", "2", "OPAQUE", "PRIVATE",
         "10", "Weekly", "", "", "2",
         "manager@company.com", "", "", "30 minutes"],
        ["Charity Run", "20/06/2026", "", "09:00", "", "", "Hyde Park, London",
         "10K charity run for NHS", "Personal", "https://charity-run.example.com", "CONFIRMED", "", "OPAQUE", "PUBLIC",
         "2880", "", "", "", "",
         "", "", "", ""],
    ]

    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    add_notes_sheet(wb)
    wb.save(r"C:\Users\simcr\OneDrive\Projects\CSV2ICS\CSV2ICS_Sample.xlsx")
    print("Sample saved")


if __name__ == "__main__":
    create_template()
    create_sample()
    print("Done!")
